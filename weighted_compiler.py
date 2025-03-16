import os
import random
import pandas as pd
from datetime import datetime, timedelta
from pydub import AudioSegment
from openpyxl import load_workbook
import time

# --- User-defined variables ---
include_paths = [
    "path/to/folder_01",
    "path/to/folder_02"
]

smoothing_constant = 5  # Smoothing factor for inverse probability weighting

# File paths
excel_file = "usage_history.xlsx"
destination_audio_folder = "/path/to/destination"
output_format = "mp3"
random_order = True

target_duration_seconds = 14400  # 4 hours
fade_in_duration = 3000  # 3 seconds
fade_out_duration = 3000  # 3 seconds

# Ensure output directory exists
os.makedirs(destination_audio_folder, exist_ok=True)

# Load music usage history
if not os.path.exists(excel_file):
    raise FileNotFoundError(f"Usage history file '{excel_file}' not found. Run the initialization script first.")

df_historic = pd.read_excel(excel_file)

# Filter available files
valid_music_df = df_historic[
    df_historic['folder_path'].isin(include_paths) & ~df_historic['deleted_renamed']
]

# Debugging: Count number of files per folder
folder_counts = {folder: df_historic[df_historic['folder_path'] == folder].shape[0] for folder in include_paths}

# Print the folder counts
print("Folder counts:")
for folder, count in folder_counts.items():
    print(f'"{folder}" -> {count} files')

# Check if valid_music_df is empty
if valid_music_df.empty:
    print("No valid music files found. Exiting script.")
    exit(0)

# Implement inverse probability weighting for selection
def get_weighted_sample(df, sample_size):
    probabilities = 1 / (df['n_usage'] + smoothing_constant)
    probabilities /= probabilities.sum()
    return df.sample(n=sample_size, weights=probabilities, replace=False)

# Function to update Excel file without overwriting formatting
def update_excel_usage(file_path, df):
    wb = load_workbook(file_path)
    ws = wb.active
    
    for index, row in df.iterrows():
        for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if cell[0].value == row['music_path']:
                cell[3].value = row['n_usage']  # Update n_usage column
                break
    
    wb.save(file_path)

# Get the start and end date for audio generation
start_date = datetime.strptime("2025-04-01", "%Y-%m-%d")
end_date = datetime.strptime("2025-04-10", "%Y-%m-%d")

# Loop through each date to generate daily compilations
current_date = start_date
while current_date <= end_date:
    current_time = time.strftime("[%Y-%m-%d %H:%M]", time.localtime())
    output_file_name = current_date.strftime("%Y-%m-%d")
    print("-"*70)
    print(f"{current_time} Generating compilation for {output_file_name}")

    # Select music files based on weighted probability
    sample_size = min(len(valid_music_df), 250)  # Adjust as needed
    selected_files = get_weighted_sample(valid_music_df, sample_size)
    
    if random_order:
        selected_files = selected_files.sample(frac=1).reset_index(drop=True)
    
    # Initialize an empty AudioSegment
    combined_audio = AudioSegment.silent(duration=0)
    audio_start_times = []
    current_position = 0
    target_duration_ms = target_duration_seconds * 1000 if target_duration_seconds else None
    
    # Process selected music files
    for _, row in selected_files.iterrows():
        audio_path = row['music_path']
        if audio_path.endswith('.wav'):
            audio = AudioSegment.from_wav(audio_path)
        elif audio_path.endswith('.mp3'):
            audio = AudioSegment.from_mp3(audio_path)
        else:
            continue
        
        audio = audio.fade_in(fade_in_duration).fade_out(fade_out_duration)
        
        # Stop if adding this track exceeds target duration
        if target_duration_ms and current_position + len(audio) > target_duration_ms + 120000:  # Allow 2-minute overflow
            print("Target duration reached. Compiling final audio...")
            break
        else:
            print("Adding:", audio_path)
        
        combined_audio += audio
        start_time = f"{current_position // 3600000:02}:{(current_position % 3600000) // 60000:02}:{(current_position % 60000) // 1000:02}"
        audio_start_times.append(f"{start_time} - {audio_path}")
        
        # Update n_usage count in the dataframe
        df_historic.loc[df_historic['music_path'] == audio_path, 'n_usage'] += 1
        
        current_position += len(audio)
    
    # Export compiled audio
    output_audio_path = os.path.join(destination_audio_folder, f"{output_file_name}.{output_format}")
    combined_audio.export(output_audio_path, format=output_format)
    
    # Save tracklist
    output_txt_path = os.path.join(destination_audio_folder, f"{output_file_name}_audio_list.txt")
    with open(output_txt_path, "w") as txt_file:
        txt_file.write("\n".join(audio_start_times))
    
    print(f"Generated: {output_audio_path}")
    print(f"Tracklist saved: {output_txt_path}")
    
    # Update Excel file without overwriting formatting
    try:
        update_excel_usage(excel_file, df_historic)
        print(f"Updated usage history in Excel file: {excel_file}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")
    
    # Move to the next date
    current_date += timedelta(days=1)
